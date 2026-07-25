#!/usr/bin/env python3
"""
Pocket Ledger Migration Script
================================
Reads DailyBudgetBackup2.sqlite and finances.xlsx and produces migration-data.json
ready to import into the Pocket Ledger PWA.

Usage:
    python3 extract.py [sqlite_path] [xlsx_path]

Defaults:
    sqlite_path = ../DailyBudgetBackup2.sqlite
    xlsx_path   = ../finances.xlsx

Output:
    migration-data.json (in current directory)
"""

import sys
import json
import sqlite3
import os
from datetime import datetime, timedelta, date

# Apple Core Data epoch: seconds since 2001-01-01
APPLE_EPOCH = datetime(2001, 1, 1)

def apple_ts(v):
    if v is None:
        return None
    try:
        dt = APPLE_EPOCH + timedelta(seconds=v)
        # Clamp dates beyond 4001 to the sentinel open-end value
        if dt.year > 4001:
            return '4001-01-01'
        return dt.strftime('%Y-%m-%d')
    except:
        return None

# Category mapping: legacy SQLite pk -> new app id
# Based on scope doc and analysis of actual data
CATEGORY_MAP = {
    1: 2,   # Food/house shop
    2: 2,   # Groceries (archived) -> Food/house shop
    3: 15,  # Socializing
    4: 16,  # Health -> archived Health
    5: 3,   # Food&Dining -> Food out (unplanned)
    6: 14,  # Sale/Expenses -> Expenses reimbursement
    7: 18,  # PersonalCare
    8: 9,   # Taxes -> Household (closest)
    9: 21,  # Investment
    10: 6,  # Takeaway
    11: 19, # Coffee -> Drinks
    12: 7,  # Charity/gifts
    13: 24, # Hobby
    14: 3,  # Food out (unplanned)
    15: 10, # Clothing
    16: 9,  # Household
    17: 22, # Fees -> General
    18: 22, # Children -> General
    19: 1,  # Travel -> Transportation
    20: 17, # Bills
    21: 20, # Sale
    22: 16, # Health (archived)
    23: 5,  # Entertainment
    24: 1,  # Transportation
    25: 12, # Education
    26: 19, # Drinks
    27: 13, # ExtraIncome
    28: 22, # General
    29: 23, # Fuel
    30: 13, # Bonus -> Extra income
    31: 22, # Pet -> General
    32: 11, # Leisure
    33: 8,  # Health and beauty
    34: 4,  # Restaurant
    35: 22, # Cigarettes -> General
    36: 7,  # Gifts -> Charity/gifts
}

# Account name -> id mapping
ACCOUNT_MAP = {
    'Current': 1, 'Current account': 1,
    'Credit cards': 2, 'Credit card': 2,
    'Vida ISA': 3, '212 ISA': 4, 'Trading 212 cash ISA': 4, 'Flex saver': 5,
    'Tembo': 6, 'SJP ISA': 7, 'SJP stocks ISA': 7,
    '212 ISA Stocks': 8, 'Trading 212 stocks ISA': 8,
    'Stu loan': 9, 'Student loan': 9,
    'APCs': 10, 'Property': 11, 'Mortgage': 12,
    'Bank of Gilulu': 13,
}

# Financial goals sheet column headers -> account ids
FG_ACCOUNT_COLS = {
    1: 1,   # Current account
    2: 2,   # Credit cards
    3: 3,   # Vida ISA
    4: 4,   # Trading 212 cash ISA
    5: 5,   # Flex saver
    6: 6,   # Tembo
    7: 13,  # Bank of Gilulu
    # col 8 = "Total cash" skip
    # col 9 = Change skip
    10: 7,  # SJP ISA
    11: 9,  # Student loan
    12: 8,  # 212 Stocks ISA
    13: 10, # APCs
    14: 11, # Property
    15: 12, # Mortgage
    # col 16 = net wealth skip
}

def migrate_sqlite(conn):
    c = conn.cursor()

    # ── Transactions ──────────────────────────────────────────────────────
    c.execute("""
        SELECT Z_PK, ZDATE, ZAMOUNT, ZCATEGORY, ZNOTE,
               ZISDAILYBUDGET, ZISEXTRAINCOME, ZISWISH, ZISEDITABLE
        FROM ZBOOKING
        WHERE ZISEDITABLE = 1
           OR (ZISEDITABLE IS NULL)
        ORDER BY ZDATE
    """)

    transactions = []
    for row in c.fetchall():
        pk, zdate, amount, cat_legacy, note, is_budget, is_extra, is_wish, is_edit = row

        date_str = apple_ts(zdate)
        if not date_str:
            continue
        if amount is None:
            continue

        # Skip daily-budget synthetic rows (ZISDAILYBUDGET = 1)
        if is_budget == 1:
            continue

        cat_id = CATEGORY_MAP.get(cat_legacy, 22)  # default General

        # Determine type
        if is_wish:
            txn_type = 'distributed_income' if amount > 0 else 'distributed_expense'
        elif amount > 0:
            txn_type = 'income'
        else:
            txn_type = 'expense'

        txn = {
            'id': pk,
            'date': date_str,
            'amount': round(amount, 2),
            'categoryId': cat_id,
            'note': (note or '').strip(),
            'type': txn_type,
            'distributionId': None,
            'createdAt': datetime.utcnow().isoformat() + 'Z',
            'updatedAt': datetime.utcnow().isoformat() + 'Z',
            'syncStatus': 'synced',
        }
        transactions.append(txn)

    # ── Recurring Expenses ────────────────────────────────────────────────
    c.execute("SELECT Z_PK, ZSTARTDATE, ZENDDATE, ZAMOUNT, ZDESC, ZTIMEOPTION FROM ZFIXEDSPENDING ORDER BY ZSTARTDATE")
    recurring_expenses = []
    now_str = date.today().isoformat()

    for row in c.fetchall():
        pk, start_ts, end_ts, amount, desc, freq_raw = row
        start = apple_ts(start_ts)
        end = apple_ts(end_ts)
        if not start or not desc or amount is None:
            continue

        # Normalise frequency
        freq_raw = (freq_raw or 'Monthly').strip().lower()
        if 'year' in freq_raw or 'annual' in freq_raw:
            freq = 'yearly'
        elif 'quarter' in freq_raw:
            freq = 'quarterly'
        elif 'week' in freq_raw:
            freq = 'monthly'  # approximate
        else:
            freq = 'monthly'

        # Active = no end date, or end date in the future (4001 = open-ended)
        is_open = (not end or end.startswith('4001') or end > now_str)
        is_active = is_open

        rec = {
            'id': pk,
            'description': desc.strip(),
            'amount': round(abs(amount or 0), 2),
            'frequency': freq,
            'startDate': start,
            'endDate': end if end and not end.startswith('4001') else '4001-01-01',
            'isShared': False,
            'sharePercent': 50,
            'category': None,
            'nextReviewDate': None,
            'isActive': bool(is_active),
        }
        recurring_expenses.append(rec)

    # Mark shared bills from scope document
    SHARED_DESCS = ['mortgage', 'council tax', 'ground rent', 'electricity', 'octopus', 'gas', 'water', 'internet', 'netflix', 'disney', 'tv licence', 'help to buy', 'htb']
    for r in recurring_expenses:
        desc_lower = r['description'].lower()
        if any(s in desc_lower for s in SHARED_DESCS):
            r['isShared'] = True

    # ── Recurring Income ──────────────────────────────────────────────────
    c.execute("SELECT Z_PK, ZSTARTDATE, ZENDDATE, ZAMOUNT, ZUSERDESCRIPTION, ZINTERVAL FROM ZREGULARINCOME ORDER BY ZSTARTDATE")
    recurring_income = []

    for row in c.fetchall():
        pk, start_ts, end_ts, amount, desc, interval = row
        start = apple_ts(start_ts)
        end = apple_ts(end_ts)
        if not start or amount is None:
            continue

        is_open = (not end or end.startswith('4001') or end > now_str)

        inc = {
            'id': pk,
            'description': (desc or 'Income').strip(),
            'amount': round(abs(amount), 2),
            'intervalDays': int(interval or 30),
            'startDate': start,
            'endDate': end if end and not end.startswith('4001') else '4001-01-01',
            'isActive': bool(is_open),
        }
        recurring_income.append(inc)

    # ── Savings Targets ───────────────────────────────────────────────────
    c.execute("SELECT Z_PK, ZSTARTDATE, ZENDDATE, ZAMOUNT, ZPERCENTAGE FROM ZDB2CYCLESAVING ORDER BY ZSTARTDATE")
    savings_targets = []

    for row in c.fetchall():
        pk, start_ts, end_ts, amount, pct = row
        start = apple_ts(start_ts)
        end = apple_ts(end_ts)
        if not start:
            continue

        target = {
            'id': pk,
            'amount': round(amount or 0, 2),
            'percentage': round(pct or 0, 4),
            'startDate': start,
            'endDate': end if end and not end.startswith('4001') else '4001-01-01',
            'description': f'Savings {start[:7]}',
        }
        savings_targets.append(target)

    # ── Distributions (ZWISH) ─────────────────────────────────────────────
    c.execute("SELECT Z_PK, ZSTARTDATE, ZENDDATE, ZAMOUNT, ZDESC, ZCATEGORY, ZISFINISHED FROM ZWISH ORDER BY ZSTARTDATE")
    distributions = []

    for row in c.fetchall():
        pk, start_ts, end_ts, amount, desc, cat_legacy, finished = row
        start = apple_ts(start_ts)
        end = apple_ts(end_ts)
        if not start or not end or amount is None:
            continue

        cat_id = CATEGORY_MAP.get(cat_legacy, 22)

        dist = {
            'id': pk,
            'description': (desc or 'Expense').strip(),
            'totalAmount': round(abs(amount), 2),
            'startDate': start,
            'endDate': end,
            'categoryId': cat_id,
            'isIncome': amount > 0,
            'isFinished': bool(finished),
        }
        distributions.append(dist)

    return transactions, recurring_expenses, recurring_income, savings_targets, distributions


def migrate_xlsx(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return [], [], []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    account_snapshots = []
    friend_holdings = []
    friend_transactions = []
    snap_id = 1
    ft_id = 1

    # ── Financial goals sheet - account snapshots ─────────────────────────
    if 'Financial goals' in wb.sheetnames:
        ws = wb['Financial goals']
        # Row 9 is headers, rows 10+ are data
        # Cols (0-indexed): 0=date, 1=Current, 2=CreditCards, 3=VidaISA, 4=212ISA,
        #   5=FlexSaver, 6=Tembo, 7=BankOfGilulu, 8=TotalCash(skip), 9=Change(skip),
        #   10=SJP, 11=StuLoan, 12=212Stocks, 13=APCs, 14=Property, 15=Mortgage

        # Map column index (0-based) to account id
        col_account = {
            0: None,   # date
            1: 1,      # Current account
            2: 2,      # Credit cards
            3: 3,      # Vida ISA
            4: 4,      # 212 cash ISA
            5: 5,      # Flex saver
            6: 6,      # Tembo
            7: 13,     # Bank of Gilulu
            8: None,   # Total cash - skip
            9: None,   # Change - skip
            10: 7,     # SJP ISA
            11: 9,     # Student loan
            12: 8,     # 212 Stocks ISA
            13: 10,    # APCs
            14: 11,    # Property
            15: 12,    # Mortgage
        }

        for row in ws.iter_rows(min_row=10, values_only=True):
            if not row[0]:
                continue
            dt = row[0]
            if isinstance(dt, datetime):
                date_str = dt.strftime('%Y-%m-%d')
            elif isinstance(dt, str):
                date_str = dt[:10]
            else:
                continue

            for col_idx, acc_id in col_account.items():
                if acc_id is None:
                    continue
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None or not isinstance(val, (int, float)):
                    continue

                account_snapshots.append({
                    'id': snap_id,
                    'accountId': acc_id,
                    'date': date_str,
                    'balance': round(float(val), 2),
                    'note': '',
                })
                snap_id += 1

    # ── Bank of Gilulu sheet ──────────────────────────────────────────────
    if 'Bank of Gilulu' in wb.sheetnames:
        ws = wb['Bank of Gilulu']

        # Dom: columns A-G (0-6), transactions from row 4 onward
        # Jesse: columns J-P (9-15), transactions from row 4 onward

        # Create holdings
        friend_holdings.append({'id': 1, 'friendName': 'Dom', 'isActive': True, 'interestRate': 0.04, 'notes': 'Bank of Gilulu'})
        friend_holdings.append({'id': 2, 'friendName': 'Jesse', 'isActive': True, 'interestRate': 0.04, 'notes': 'Bank of Gilulu'})

        for row in ws.iter_rows(min_row=4, values_only=True):
            # Dom: col C (index 2) = amount, col D (index 3) = date
            if len(row) > 3:
                dom_amount = row[2]
                dom_date = row[3]
                if isinstance(dom_amount, (int, float)) and dom_amount != 0 and isinstance(dom_date, datetime):
                    friend_transactions.append({
                        'id': ft_id,
                        'holdingId': 1,
                        'date': dom_date.strftime('%Y-%m-%d'),
                        'amount': round(float(dom_amount), 2),
                        'note': '',
                    })
                    ft_id += 1

            # Jesse: col K (index 10) = amount, col L (index 11) = date
            if len(row) > 11:
                jesse_amount = row[10]
                jesse_date = row[11]
                if isinstance(jesse_amount, (int, float)) and jesse_amount != 0 and isinstance(jesse_date, datetime):
                    friend_transactions.append({
                        'id': ft_id,
                        'holdingId': 2,
                        'date': jesse_date.strftime('%Y-%m-%d'),
                        'amount': round(float(jesse_amount), 2),
                        'note': '',
                    })
                    ft_id += 1

    return account_snapshots, friend_holdings, friend_transactions


def main():
    sqlite_path = sys.argv[1] if len(sys.argv) > 1 else '../DailyBudgetBackup2.sqlite'
    xlsx_path   = sys.argv[2] if len(sys.argv) > 2 else '../finances.xlsx'

    if not os.path.exists(sqlite_path):
        print(f"SQLite file not found: {sqlite_path}")
        print("Usage: python3 extract.py <sqlite_path> <xlsx_path>")
        sys.exit(1)

    print(f"Reading SQLite: {sqlite_path}")
    conn = sqlite3.connect(sqlite_path)
    transactions, recurring_expenses, recurring_income, savings_targets, distributions = migrate_sqlite(conn)
    conn.close()

    account_snapshots, friend_holdings, friend_transactions = [], [], []
    if os.path.exists(xlsx_path):
        print(f"Reading Excel: {xlsx_path}")
        account_snapshots, friend_holdings, friend_transactions = migrate_xlsx(xlsx_path)
    else:
        print(f"Excel file not found: {xlsx_path} (skipping account snapshots)")

    # Use the standard seed categories from the app (don't override)
    output = {
        'transactions':      transactions,
        'recurringExpenses': recurring_expenses,
        'recurringIncome':   recurring_income,
        'savingsTargets':    savings_targets,
        'distributions':     distributions,
        'accountSnapshots':  account_snapshots,
        'friendHoldings':    friend_holdings,
        'friendTransactions':friend_transactions,
        'exportedAt': datetime.utcnow().isoformat() + 'Z',
    }

    out_path = 'migration-data.json'
    with open(out_path, 'w') as f:
        json.dump(output, f, indent=2)

    print(f"\nMigration complete: {out_path}")
    print(f"  {len(transactions)} transactions")
    print(f"  {len(recurring_expenses)} recurring expenses")
    print(f"  {len(recurring_income)} recurring income records")
    print(f"  {len(savings_targets)} savings targets")
    print(f"  {len(distributions)} distributions (big expenses)")
    print(f"  {len(account_snapshots)} account snapshots")
    print(f"  {len(friend_holdings)} friend holdings")
    print(f"  {len(friend_transactions)} friend transactions")
    print(f"\nImport this file in the app: Settings > Import data")


if __name__ == '__main__':
    main()
